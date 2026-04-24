from __future__ import annotations

import argparse
import csv
import gzip
import re
import shutil
import zipfile
from pathlib import Path
from typing import Iterable


ROOT = Path(".")
DEFAULT_INPUT_DIR = ROOT / "data" / "lgd_villages_input"
OUT_DIR = ROOT / "data"
PKG_DATA = ROOT / "indic_places" / "data"

CUSTOM_PLACES = PKG_DATA / "custom_places.txt"
OUT_UNIQUE = OUT_DIR / "all_india_villages_unique.txt"
OUT_FULL = OUT_DIR / "all_india_villages_full.csv"
OUT_FULL_GZ = OUT_DIR / "all_india_villages_full.csv.gz"

PYPROJECT = ROOT / "pyproject.toml"
README = ROOT / "README.md"

VILLAGE_HINTS = (
    "village name",
    "village_name",
    "villagename",
    "village english",
    "village_english",
    "village",
    "villages",
    "local body name",
    "localbodyname",
)

STATE_HINTS = (
    "state name",
    "state_name",
    "statename",
    "state",
)

DISTRICT_HINTS = (
    "district name",
    "district_name",
    "districtname",
    "district",
)

SUBDISTRICT_HINTS = (
    "sub district",
    "subdistrict",
    "sub-district",
    "tehsil",
    "taluk",
    "taluka",
    "block",
)

PIN_HINTS = (
    "pin",
    "pincode",
    "pin code",
    "postal code",
    "postal",
)

LGD_HINTS = (
    "village code",
    "village_code",
    "villagecode",
    "lgd",
    "lgd code",
    "local government directory code",
)


def clean_cell(value: object) -> str:
    s = str(value or "").replace("\ufeff", "").strip()
    s = re.sub(r"\s+", " ", s)
    return s.strip(" ,;:-|[]{}()")


def name_key(value: str) -> str:
    """Duplicate-safe key: same spelling with spaces/punctuation/case differences becomes same key."""
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def safe_extract_filename(name: str) -> str:
    """Make ZIP member names safe for Windows file extraction."""
    base = Path(str(name).replace("\\", "/")).name
    base = re.sub(r'[<>:"/\\|?*]+', "_", base)
    base = re.sub(r"\s+", " ", base).strip(" ._")
    return base or "extracted_file"


def pretty_name(value: str) -> str:
    s = clean_cell(value)

    if not s:
        return ""

    if re.fullmatch(r"[\d.\-_/]+", s):
        return ""

    if name_key(s) in {"VILLAGE", "VILLAGENAME", "NAME", "NA", "NONE", "NULL", "NIL", "NAN"}:
        return ""

    # Remove repeated header/noise suffix if present.
    s = re.sub(r"(?i)\b(?:village|gram|lgd|code)\s*[:=-]?\s*$", "", s).strip()

    # Preserve mixed case, title-case full uppercase rows.
    if s.isupper() and len(s) > 3:
        s = s.title()

    return s


def col_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def find_col(fieldnames: list[str], hints: tuple[str, ...]) -> str | None:
    cols = [(c, col_key(c)) for c in fieldnames]

    for col, ck in cols:
        for hint in hints:
            if ck == col_key(hint):
                return col

    for col, ck in cols:
        for hint in hints:
            hk = col_key(hint)
            if hk and hk in ck:
                return col

    return None


def detect_delimiter(sample: str) -> str:
    candidates = [",", "\t", "|", ";"]
    return max(candidates, key=lambda c: sample.count(c))


def iter_csv_rows(path: Path) -> Iterable[dict]:
    raw = path.read_bytes()
    text = None

    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            text = raw.decode(enc)
            break
        except Exception:
            continue

    if text is None:
        return

    delimiter = detect_delimiter(text[:4096])
    reader = csv.DictReader(text.splitlines(), delimiter=delimiter)

    for row in reader:
        yield row


def iter_excel_rows_pandas(path: Path) -> Iterable[dict]:
    # Read .xls/.xlsx using pandas. Required for LGD Excel 97-2003 .xls files.
    # Install if needed: python -m pip install pandas xlrd openpyxl
    try:
        import pandas as pd
    except Exception:
        print(f"SKIP Excel because pandas is not installed: {path}")
        print("Install: python -m pip install pandas xlrd openpyxl")
        return

    try:
        sheets = pd.read_excel(path, sheet_name=None, dtype=str)
    except Exception as exc:
        print(f"SKIP Excel because it could not be read: {path}")
        print(f"Reason: {exc}")
        print("For .xls install: python -m pip install xlrd")
        return

    for _sheet_name, df in sheets.items():
        df = df.fillna("")
        columns = [str(c).strip() for c in df.columns]

        for _, row in df.iterrows():
            yield {
                columns[i]: clean_cell(row.iloc[i])
                for i in range(len(columns))
            }


def iter_xlsx_rows(path: Path) -> Iterable[dict]:
    try:
        import openpyxl
    except Exception:
        print(f"SKIP XLSX because openpyxl is not installed: {path}")
        return

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    for ws in wb.worksheets:
        header = None

        for raw_row in ws.iter_rows(values_only=True):
            values = [clean_cell(v) for v in raw_row]

            if not any(values):
                continue

            if header is None:
                header = values
                continue

            yield {header[i]: values[i] if i < len(values) else "" for i in range(len(header))}


def unzip_supported(path: Path, temp_dir: Path) -> list[Path]:
    out = []

    if not zipfile.is_zipfile(path):
        return out

    temp_dir.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(path) as zf:
        for name in zf.namelist():
            if name.endswith("/"):
                continue

            suffix = Path(name).suffix.lower()
            if suffix not in {".csv", ".txt", ".tsv", ".xlsx", ".xls"}:
                continue

            safe_name = safe_extract_filename(name)
            target = temp_dir / safe_name

            # Avoid overwriting if ZIP contains repeated unsafe names.
            if target.exists():
                stem = target.stem
                suffix = target.suffix
                i = 2
                while True:
                    alt = temp_dir / f"{stem}_{i}{suffix}"
                    if not alt.exists():
                        target = alt
                        break
                    i += 1

            target.write_bytes(zf.read(name))
            out.append(target)

    return out


def gunzip_supported(path: Path, temp_dir: Path) -> list[Path]:
    temp_dir.mkdir(parents=True, exist_ok=True)
    target = temp_dir / path.with_suffix("").name

    with gzip.open(path, "rb") as src:
        target.write_bytes(src.read())

    return [target]


def iter_rows(path: Path, temp_dir: Path) -> Iterable[dict]:
    suffix = path.suffix.lower()

    if suffix == ".zip":
        for child in unzip_supported(path, temp_dir):
            yield from iter_rows(child, temp_dir)
        return

    if suffix == ".gz":
        for child in gunzip_supported(path, temp_dir):
            yield from iter_rows(child, temp_dir)
        return

    if suffix == ".xls":
        yield from iter_excel_rows_pandas(path)
        return

    if suffix == ".xlsx":
        try:
            yield from iter_xlsx_rows(path)
        except Exception:
            yield from iter_excel_rows_pandas(path)
        return

    if suffix in {".csv", ".txt", ".tsv"}:
        yield from iter_csv_rows(path)
        return


def input_files(input_dir: Path) -> list[Path]:
    if not input_dir.exists():
        return []

    return [
        p for p in input_dir.rglob("*")
        if p.is_file() and p.suffix.lower() in {".csv", ".txt", ".tsv", ".xlsx", ".xls", ".zip", ".gz"}
    ]


def extract_village_record(row: dict) -> dict | None:
    fieldnames = list(row.keys())

    village_col = find_col(fieldnames, VILLAGE_HINTS)
    if not village_col:
        return None

    state_col = find_col(fieldnames, STATE_HINTS)
    district_col = find_col(fieldnames, DISTRICT_HINTS)
    subdistrict_col = find_col(fieldnames, SUBDISTRICT_HINTS)
    pin_col = find_col(fieldnames, PIN_HINTS)
    lgd_col = find_col(fieldnames, LGD_HINTS)

    village = pretty_name(row.get(village_col, ""))
    if not village:
        return None

    return {
        "village": village,
        "state": clean_cell(row.get(state_col, "")) if state_col else "",
        "district": clean_cell(row.get(district_col, "")) if district_col else "",
        "subdistrict": clean_cell(row.get(subdistrict_col, "")) if subdistrict_col else "",
        "pincode": clean_cell(row.get(pin_col, "")) if pin_col else "",
        "lgd_code": clean_cell(row.get(lgd_col, "")) if lgd_col else "",
    }


def read_existing_custom_places(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}

    out = {}

    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        name = line.split("#", 1)[0].strip()
        key = name_key(name)

        if key:
            out.setdefault(key, name)

    return out


def update_pyproject(version: str) -> None:
    if not PYPROJECT.exists():
        return

    text = PYPROJECT.read_text(encoding="utf-8-sig")
    text = re.sub(r'(?m)^version\s*=.*$', f'version = "{version}"', text)
    text = re.sub(
        r'(?m)^description\s*=.*$',
        'description = "Indian place-name lookup, OCR address cleanup, all-India village vocabulary, correction, extraction, and address intelligence"',
        text,
    )
    PYPROJECT.write_text(text, encoding="utf-8")


def update_readme() -> None:
    if not README.exists():
        return

    note = """
### All-India village vocabulary

The package can include village names imported from official/LGD-style village datasets.

The import script adds only unique village names to:

```text
indic_places/data/custom_places.txt
```

Duplicate protection uses a normalized key, so names already present with different case, spacing, or punctuation are not added again.

This improves village-level matching and correction, for example:

```python
ip.correct_place_name("DORA CHHAPR")
ip.correct_place_name("MOHAN CHHAPR")
```
"""

    text = README.read_text(encoding="utf-8", errors="ignore")
    if "### All-India village vocabulary" not in text:
        README.write_text(text.rstrip() + "\n\n" + note.strip() + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", default=str(DEFAULT_INPUT_DIR))
    parser.add_argument("--version", default="1.3.4")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not (ROOT / "indic_places").exists():
        raise SystemExit("ERROR: Run this from indic_names_library repo root.")

    input_dir = Path(args.input_dir)
    input_dir.mkdir(parents=True, exist_ok=True)

    files = input_files(input_dir)
    if not files:
        print("No village dataset files found.")
        print(f"Put LGD/all-India village CSV/XLSX/ZIP files here:")
        print(input_dir)
        print()
        print("Then run again:")
        print("python import_all_india_villages_unique.py")
        return

    temp_dir = OUT_DIR / "_tmp_village_import"
    if temp_dir.exists():
        shutil.rmtree(temp_dir)
    temp_dir.mkdir(parents=True, exist_ok=True)

    structured_records = []
    structured_seen = set()

    for file in files:
        print(f"Reading: {file}")
        before = len(structured_records)

        for row in iter_rows(file, temp_dir):
            rec = extract_village_record(row)
            if not rec:
                continue

            structured_key = (
                name_key(rec["village"]),
                name_key(rec["state"]),
                name_key(rec["district"]),
                name_key(rec["subdistrict"]),
                rec["pincode"],
                rec["lgd_code"],
            )

            if structured_key in structured_seen:
                continue

            structured_seen.add(structured_key)
            structured_records.append(rec)

        print(f"  records added: {len(structured_records) - before:,}")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    PKG_DATA.mkdir(parents=True, exist_ok=True)

    unique_village_names: dict[str, str] = {}
    for rec in structured_records:
        key = name_key(rec["village"])
        if key:
            unique_village_names.setdefault(key, rec["village"])

    existing_custom = read_existing_custom_places(CUSTOM_PLACES)

    new_only = {
        key: village
        for key, village in unique_village_names.items()
        if key not in existing_custom
    }

    merged_custom = dict(existing_custom)
    merged_custom.update(new_only)

    print("=" * 80)
    print("SUMMARY")
    print(f"Structured village records read : {len(structured_records):,}")
    print(f"Unique village names found      : {len(unique_village_names):,}")
    print(f"Already present in custom_places: {len(unique_village_names) - len(new_only):,}")
    print(f"New unique names to add         : {len(new_only):,}")
    print(f"Final custom_places count       : {len(merged_custom):,}")

    if args.dry_run:
        print("DRY RUN ONLY. No files changed.")
        return

    with OUT_FULL.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["village", "state", "district", "subdistrict", "pincode", "lgd_code"],
        )
        writer.writeheader()
        writer.writerows(structured_records)

    with OUT_FULL.open("rb") as src, gzip.open(OUT_FULL_GZ, "wb", compresslevel=9) as dst:
        shutil.copyfileobj(src, dst)

    OUT_UNIQUE.write_text(
        "\n".join(sorted(unique_village_names.values(), key=lambda x: x.upper())) + "\n",
        encoding="utf-8",
    )

    CUSTOM_PLACES.write_text(
        "\n".join(sorted(merged_custom.values(), key=lambda x: x.upper())) + "\n",
        encoding="utf-8",
    )

    update_pyproject(args.version)
    update_readme()

    print("=" * 80)
    print("DONE")
    print(f"Wrote structured CSV     : {OUT_FULL}")
    print(f"Wrote compressed CSV     : {OUT_FULL_GZ}")
    print(f"Wrote unique village list: {OUT_UNIQUE}")
    print(f"Updated custom places    : {CUSTOM_PLACES}")
    print(f"Version set to           : {args.version}")


if __name__ == "__main__":
    main()
